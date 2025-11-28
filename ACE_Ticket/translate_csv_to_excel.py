#!/usr/bin/env python3
"""
将 CSV 文件中的 workflow_data 和 learned_strategies 列翻译为中文，并保存为 Excel 文件
"""
import pandas as pd
import json
import asyncio
from openai import AsyncOpenAI
import logging
import os
from typing import Optional
from dotenv import load_dotenv

# 加载环境变量
load_dotenv()
# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)



api_key = os.getenv("OPENAI_API_KEY")
# 初始化 OpenAI 客户端
def get_openai_client():
    """获取 OpenAI 客户端"""
    # 优先使用平台密钥和网关
    return AsyncOpenAI(
        api_key=api_key,
    )

async def translate_text(text: str, client: AsyncOpenAI, max_retries: int = 3) -> str:
    """
    使用 OpenAI API 将文本翻译为中文
    
    Args:
        text: 要翻译的文本
        client: OpenAI 客户端
        max_retries: 最大重试次数
        
    Returns:
        翻译后的中文文本
    """
    if not text or pd.isna(text):
        return ""
    
    # 如果文本已经是中文（包含中文字符），直接返回
    if any('\u4e00' <= char <= '\u9fff' for char in str(text)):
        return str(text)
    
    prompt = f"""请将以下内容翻译为中文，保持原有的格式和结构（如 JSON 格式、换行等），不要遗漏任何内容。只返回翻译后的内容，不要添加任何解释或说明。

原文：
```json
{text}
```
翻译："""
    
    for attempt in range(max_retries):
        try:
            response = await client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "你是一个专业的翻译助手，擅长将英文内容准确翻译为中文，保持原有格式。"},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=4000,
                temperature=0.3
            )
            
            translated = response.choices[0].message.content.strip()
            logger.info(f"翻译成功，原文长度: {len(str(text))}, 译文长度: {len(translated)}")
            return translated
            
        except Exception as e:
            logger.warning(f"翻译失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
            if attempt == max_retries - 1:
                logger.error(f"翻译最终失败，返回原文: {str(e)}")
                return str(text)
            await asyncio.sleep(2 ** attempt)  # 指数退避

async def translate_workflow_data(workflow_data: str, client: AsyncOpenAI) -> str:
    """
    翻译 workflow_data 列（JSON 格式）
    
    Args:
        workflow_data: workflow_data 列的原始内容
        client: OpenAI 客户端
        
    Returns:
        翻译后的格式化 JSON 字符串
    """
    if not workflow_data or pd.isna(workflow_data):
        return ""
    
    try:
        # 尝试解析 JSON
        data = json.loads(workflow_data)
        
        # 检查是否包含 next_steps 字段
        has_next_steps = isinstance(data, dict) and 'next_steps' in data
        
        # 将 JSON 转换为格式化的字符串进行翻译
        formatted_json = json.dumps(data, ensure_ascii=False, indent=2)
        
        # 使用更明确的提示词，特别强调要翻译数组中的内容
        prompt = f"""请将以下 JSON 内容完整翻译为中文，包括所有字段和数组中的每一项内容。特别注意：
1. 必须翻译 next_steps 数组中的每一个步骤
2. 保持 JSON 格式和结构不变
3. 只翻译文本内容，不要改变 JSON 的键名（除非键名本身是英文需要翻译）
4. 不要遗漏任何内容
5. 只返回翻译后的 JSON，不要添加任何解释

原文：
```json
{formatted_json}
```

翻译后的 JSON："""
        
        # 翻译整个 JSON
        translated = ""
        for attempt in range(3):
            try:
                response = await client.chat.completions.create(
                    model= "gpt-4o",
                    messages=[
                        {"role": "system", "content": "你是一个专业的 JSON 翻译助手，擅长将 JSON 中的英文内容准确翻译为中文，包括数组中的所有项，保持 JSON 格式完整。"},
                        {"role": "user", "content": prompt}
                    ],
                    max_tokens=4000,
                    temperature=0.3
                )
                translated = response.choices[0].message.content.strip()
                break
            except Exception as e:
                logger.warning(f"翻译失败 (尝试 {attempt + 1}/3): {str(e)}")
                if attempt == 2:
                    translated = formatted_json
                await asyncio.sleep(2 ** attempt)
        
        # 尝试将翻译后的内容解析回 JSON 并格式化
        try:
            # 清理可能的代码块标记
            import re
            # 移除可能的 ```json 和 ``` 标记
            translated = re.sub(r'```json\s*', '', translated)
            translated = re.sub(r'```\s*$', '', translated, flags=re.MULTILINE)
            translated = translated.strip()
            
            # 尝试提取 JSON 对象（如果 AI 添加了说明文字）
            json_match = re.search(r'\{.*\}', translated, re.DOTALL)
            if json_match:
                translated = json_match.group()
            
            # 尝试解析翻译后的 JSON
            translated_data = json.loads(translated)
            
            # 验证 next_steps 是否被翻译（如果原数据有 next_steps）
            if has_next_steps and isinstance(translated_data, dict) and 'next_steps' in translated_data:
                original_steps = data.get('next_steps', [])
                translated_steps = translated_data.get('next_steps', [])
                
                # 检查是否有未翻译的步骤（包含大量英文字符）
                if isinstance(original_steps, list) and isinstance(translated_steps, list):
                    for i, (orig, trans) in enumerate(zip(original_steps, translated_steps)):
                        # 如果翻译后的步骤仍然包含大量英文，尝试单独翻译
                        if isinstance(trans, str) and isinstance(orig, str):
                            # 检查是否主要是英文（超过50%的字符是英文）
                            english_chars = sum(1 for c in trans if c.isascii() and c.isalpha())
                            total_chars = len([c for c in trans if c.isalnum()])
                            if total_chars > 0 and english_chars / total_chars > 0.5:
                                # 单独翻译这个步骤
                                try:
                                    step_translated = await translate_text(orig, client)
                                    translated_steps[i] = step_translated
                                    logger.info(f"单独翻译了 next_steps[{i}]")
                                except:
                                    pass
                    
                    translated_data['next_steps'] = translated_steps
            
            # 格式化输出，确保中文正确显示
            formatted_result = json.dumps(translated_data, ensure_ascii=False, indent=2)
            return formatted_result
            
        except json.JSONDecodeError as e:
            logger.warning(f"解析翻译后的 JSON 失败: {str(e)}")
            # 如果解析失败，尝试提取 JSON 部分
            import re
            json_match = re.search(r'\{.*\}', translated, re.DOTALL)
            if json_match:
                try:
                    extracted_json = json.loads(json_match.group())
                    return json.dumps(extracted_json, ensure_ascii=False, indent=2)
                except:
                    pass
            # 如果无法提取，返回翻译后的内容（保持原格式）
            return translated
    except json.JSONDecodeError:
        # 如果不是有效的 JSON，直接翻译
        return await translate_text(workflow_data, client)

async def translate_learned_strategies(learned_strategies: str, client: AsyncOpenAI) -> str:
    """
    翻译 learned_strategies 列
    
    Args:
        learned_strategies: learned_strategies 列的原始内容
        client: OpenAI 客户端
        
    Returns:
        翻译后的内容
    """
    return await translate_text(learned_strategies, client)

async def process_csv_to_excel(input_csv: str, output_excel: str, batch_size: int = 10):
    """
    处理 CSV 文件，翻译指定列并保存为 Excel
    
    Args:
        input_csv: 输入 CSV 文件路径
        output_excel: 输出 Excel 文件路径
        batch_size: 批处理大小（每次处理的记录数）
    """
    logger.info(f"开始读取 CSV 文件: {input_csv}")
    
    # 读取 CSV 文件
    df = pd.read_csv(input_csv)
    logger.info(f"成功读取 {len(df)} 条记录")
    
    # 检查必要的列是否存在
    # if 'workflow_data' not in df.columns:
    #     raise ValueError("CSV 文件中缺少 'workflow_data' 列")
    if 'learned_strategies' not in df.columns:
        raise ValueError("CSV 文件中缺少 'learned_strategies' 列")
    
    # 初始化 OpenAI 客户端
    client = get_openai_client()
    logger.info("OpenAI 客户端初始化成功")
    
    # 创建新列用于存储翻译结果
    # df['workflow_data_zh'] = ""
    df['learned_strategies_zh'] = ""
    
    # 分批处理
    total_rows = len(df)
    logger.info(f"开始翻译，共 {total_rows} 条记录，批处理大小: {batch_size}")
    
    for i in range(0, total_rows, batch_size):
        batch_end = min(i + batch_size, total_rows)
        logger.info(f"处理第 {i+1} 到 {batch_end} 条记录...")
        
        # 创建翻译任务
        tasks_workflow = []
        tasks_strategies = []
        indices = []
        
        for idx in range(i, batch_end):
            row = df.iloc[idx]
            indices.append(idx)
            
            # 翻译 workflow_data
            # if pd.notna(row['workflow_data']) and str(row['workflow_data']).strip():
            #     tasks_workflow.append(translate_workflow_data(str(row['workflow_data']), client))
            # else:
            #     tasks_workflow.append(asyncio.create_task(asyncio.sleep(0, result="")))
            
            # 翻译 learned_strategies
            if pd.notna(row['learned_strategies']) and str(row['learned_strategies']).strip():
                tasks_strategies.append(translate_learned_strategies(str(row['learned_strategies']), client))
            else:
                tasks_strategies.append(asyncio.create_task(asyncio.sleep(0, result="")))
        
        # 并发执行翻译任务
        try:
            # results_workflow = await asyncio.gather(*tasks_workflow, return_exceptions=True)
            results_strategies = await asyncio.gather(*tasks_strategies, return_exceptions=True)
            
            # 将结果写入 DataFrame
            for j, idx in enumerate(indices):
                # 处理 workflow_data 翻译结果
                # if isinstance(results_workflow[j], Exception):
                #     logger.error(f"翻译第 {idx+1} 行的 workflow_data 失败: {str(results_workflow[j])}")
                #     df.at[idx, 'workflow_data_zh'] = str(df.iloc[idx]['workflow_data']) if pd.notna(df.iloc[idx]['workflow_data']) else ""
                # else:
                #     df.at[idx, 'workflow_data_zh'] = results_workflow[j]
                
                # 处理 learned_strategies 翻译结果
                if isinstance(results_strategies[j], Exception):
                    logger.error(f"翻译第 {idx+1} 行的 learned_strategies 失败: {str(results_strategies[j])}")
                    df.at[idx, 'learned_strategies_zh'] = str(df.iloc[idx]['learned_strategies']) if pd.notna(df.iloc[idx]['learned_strategies']) else ""
                else:
                    df.at[idx, 'learned_strategies_zh'] = results_strategies[j]
                    
        except Exception as e:
            logger.error(f"批处理翻译失败: {str(e)}")
            raise
        
        logger.info(f"已完成 {batch_end}/{total_rows} 条记录")
    
    # 保存为 Excel 文件
    logger.info(f"开始保存 Excel 文件: {output_excel}")
    
    # 使用 ExcelWriter 以便设置列宽
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='翻译结果')
        
        # 获取工作表
        worksheet = writer.sheets['翻译结果']
        
        # 设置列宽，特别是 JSON 列
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
        
        for idx, col in enumerate(df.columns, start=1):
            column_letter = get_column_letter(idx)
            if col in ['workflow_data', 'workflow_data_zh']:
                # JSON 列设置更宽的宽度
                worksheet.column_dimensions[column_letter].width = 80
            elif col in ['learned_strategies', 'learned_strategies_zh']:
                # 策略列也设置较宽
                worksheet.column_dimensions[column_letter].width = 100
            else:
                # 其他列自动调整
                worksheet.column_dimensions[column_letter].width = 20
        
        # 设置文本换行，方便查看长文本（特别是翻译后的列）
        translation_cols = ['learned_strategies_zh']
        for col_name in translation_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                column_letter = get_column_letter(col_idx)
                # 为整列设置换行和对齐方式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    logger.info(f"成功保存 Excel 文件: {output_excel}")
    logger.info(f"共处理 {total_rows} 条记录")

def main():
    """主函数"""
    input_csv = "ticket_1128.csv"
    output_excel = "ticket_1128_zh.xlsx"
    
    if not os.path.exists(input_csv):
        raise FileNotFoundError(f"找不到输入文件: {input_csv}")
    
    # 检查配置
    try:
        client = get_openai_client()
        logger.info("OpenAI 客户端配置检查通过")
    except ValueError as e:
        print(f"\n❌ 错误: {e}")
        print("\n请设置以下环境变量之一：")
        print("  1. OPENAI_PLATFORM_KEY 和 AIOP_GATEWAY_URL（推荐）")
        print("  2. OPENAI_API_KEY")
        print("\n或者在 .env 文件中配置相应的值")
        return
    
    # 运行异步处理
    try:
        asyncio.run(process_csv_to_excel(input_csv, output_excel, batch_size=5))
        print(f"\n✅ 翻译完成！输出文件: {output_excel}")
    except Exception as e:
        logger.error(f"处理失败: {str(e)}")
        raise

if __name__ == "__main__":
    main()

