#!/usr/bin/env python3
"""
格式化 Excel 文件中的 JSON 列（workflow_data, ground_truth, learned_strategies）
将 JSON 内容格式化为带缩进的格式
"""
import pandas as pd
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import sys

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def format_json_string(json_str: str) -> str:
    """
    格式化 JSON 字符串
    
    Args:
        json_str: JSON 字符串
        
    Returns:
        格式化后的 JSON 字符串，如果不是有效的 JSON 则返回原字符串
    """
    if not json_str or pd.isna(json_str):
        return ""
    
    json_str = str(json_str).strip()
    if not json_str:
        return ""
    
    # 去除 ```json 和 ``` 标记
    import re
    # 去除开头的 ```json 或 ```json\n
    json_str = re.sub(r'^```json\s*\n?', '', json_str, flags=re.IGNORECASE)
    # 去除结尾的 ``` 或 \n```
    json_str = re.sub(r'\n?```\s*$', '', json_str, flags=re.IGNORECASE)
    json_str = json_str.strip()
    
    if not json_str:
        return ""
    
    try:
        # 尝试解析 JSON
        data = json.loads(json_str)
        # 格式化为带缩进的 JSON
        formatted = json.dumps(data, ensure_ascii=False, indent=2)
        return formatted
    except json.JSONDecodeError:
        # 如果不是有效的 JSON，尝试提取 JSON 部分
        # 尝试提取 JSON 对象或数组
        json_match = re.search(r'(\{.*\}|\[.*\])', json_str, re.DOTALL)
        if json_match:
            try:
                extracted_json = json.loads(json_match.group(1))
                return json.dumps(extracted_json, ensure_ascii=False, indent=2)
            except:
                pass
        # 如果无法解析，返回原字符串
        logger.warning(f"无法解析为 JSON，保持原格式: {json_str[:100]}...")
        return json_str
    except Exception as e:
        logger.error(f"格式化 JSON 时出错: {str(e)}")
        return json_str

def format_excel_json_columns(excel_path: str, output_path: str = None):
    """
    格式化 Excel 文件中的 JSON 列
    
    Args:
        excel_path: Excel 文件路径
        output_path: 输出文件路径，如果为 None 则覆盖原文件
    """
    if output_path is None:
        output_path = excel_path
    
    logger.info(f"开始读取 Excel 文件: {excel_path}")
    
    # 读取 Excel 文件
    df = pd.read_excel(excel_path, engine='openpyxl')
    logger.info(f"成功读取 {len(df)} 条记录，{len(df.columns)} 列")
    
    # 需要格式化的列
    json_columns = ['learned_strategies_zh']
    
    # 检查哪些列存在
    existing_columns = [col for col in json_columns if col in df.columns]
    if not existing_columns:
        logger.warning(f"未找到需要格式化的列: {json_columns}")
        return
    
    logger.info(f"找到需要格式化的列: {existing_columns}")
    
    # 格式化每一列
    for col in existing_columns:
        logger.info(f"开始格式化列: {col}")
        formatted_count = 0
        for idx in df.index:
            if pd.notna(df.at[idx, col]) and str(df.at[idx, col]).strip():
                original = str(df.at[idx, col])
                formatted = format_json_string(original)
                if formatted != original:
                    df.at[idx, col] = formatted
                    formatted_count += 1
        
        logger.info(f"列 {col} 格式化完成，共格式化 {formatted_count} 条记录")
    
    # 保存为 Excel 文件
    logger.info(f"开始保存 Excel 文件: {output_path}")
    
    # 使用 ExcelWriter 以便设置列宽和格式
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='翻译结果')
        
        # 获取工作表
        worksheet = writer.sheets['翻译结果']
        
        # 设置列宽和格式
        for idx, col in enumerate(df.columns, start=1):
            column_letter = get_column_letter(idx)
            
            # JSON 列设置更宽的宽度
            if col in json_columns:
                worksheet.column_dimensions[column_letter].width = 80
            elif col.endswith('_zh') and col not in json_columns:
                # 翻译后的列也设置较宽（但不在 json_columns 中的）
                worksheet.column_dimensions[column_letter].width = 100
            else:
                # 其他列自动调整
                worksheet.column_dimensions[column_letter].width = 20
        
        # 设置文本换行，方便查看长文本（特别是 JSON 列）
        for col_name in json_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                column_letter = get_column_letter(col_idx)
                # 为整列设置换行和对齐方式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    logger.info(f"成功保存 Excel 文件: {output_path}")
    logger.info(f"共处理 {len(df)} 条记录")

if __name__ == "__main__":
    # 默认处理 ace_email_learning_records_zh.xlsx
    excel_file = "ticket_1128_zh.xlsx"
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    
    output_file = None
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    try:
        format_excel_json_columns(excel_file, output_file)
        logger.info("格式化完成！")
    except Exception as e:
        logger.error(f"格式化失败: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

